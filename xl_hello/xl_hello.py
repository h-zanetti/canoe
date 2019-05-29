from openpyxl.styles import Font, PatternFill
import openpyxl as xl
import datetime as dt

font = Font()
fill = PatternFill()

wb = xl.load_workbook('test.xlsx')

ws = wb.get_sheet_by_name('Tests')

# ws['B1'] = 'Hello, world!'
# c = ws['B1']
# c.font = Font(bold=True)
# c.fill = PatternFill("solid", fgColor='FFFF99')

wb.save(filename='test.xlsx')