import datetime as dt
import calendar
from oracle import *
from openpyxl.styles import Font, PatternFill

# Styling Classes
font = Font(name='Calibri', size=12)
fill = PatternFill()
def currency_format(cell):
    cell.number_format = '"$"#,##0.00'

def invoice_font(row, ws):
    for col in range(2, 12):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name='Calibri', size=12)
        # Currency Format
        if col == 10 or col == 11:
            currency_format(cell)

# Invoice Information
def get_max_imp(imp):
    for rate in rate_card:
        if imp <= rate:
            return rate

def get_next_cpm(cpm, invoice):
    for r in range(17, 26):
        if round(invoice[f"{invoice_map['rate_cards']}{r}"].value, 2) == cpm:
            return round(invoice[f"{invoice_map['rate_cards']}{r+1}"].value, 2)

def update(id, total_imp, rate_card, invoice_num):
        dic = {
        'total_imp': total_imp,
        'rate_card': rate_card,
        'invoice_num': invoice_num,
        }

        update = f'UPDATE invoice_generator_info SET '
        for key in dic:
            update += f'{key} = {dic[key]}, '
        update = update[:-2]
        update += f" WHERE id = '{id}'"
       
        return update

backfill = {
    'programmers': ['TVONE', 'REELZ', 'KIDGENIUS', 'CROWN', 'SONY', 'KABILLION'],
    'key': ['CBFM', 'Backfill'],
    }

# Date
today = dt.date.today()
start_date = dt.date(today.year, today.month-1, 1)
end_date = dt.date(today.year, today.month-1, calendar.monthrange(today.year, today.month-1)[1])

# Excel Mapping
invoice_map = {
    # Header
    'date': 'J1', 'invoice_num': 'J2', 'period_start': 'D18', 'period_end': 'D19', 'programmer': 'D20', 'networks': 'D21', 'previous_YTD_imp': 'D22', 'rate_cards': 'I',
    # Body
    'YTD_imp': 'J', 'campaign_id': 'C', 'campaign_name': 'D', 'network': 'E', 'start_date': 'F', 'end_date': 'G', 
    'month_imp': 'H', 'cpm':  'I', 'total': 'J',
    # Foot
    'sub_networks': 'G', 'foot_networks': 'I', 'foot_total': 'F', 'amount_due': 'I',
}
df_map = {
    'campaign_id': 'A', 'campaign_name': 'B', 'network': 'C',
    'start_date': 'D', 'end_date': 'E', 'month_imp': 'F'
}

# Programmers Info
query = '''
SELECT id, abbreviation, title, total_imp, rate_card, networks, invoice_num, start_point, bill_to, attention, address, state, contact
FROM invoice_generator_info
WHERE id=18
ORDER BY TITLE
'''
cur = cursor.execute(query)
programmers = cur.fetchall()
db_map = {
    'id': 0,
    'abbreviation': 1,
    'title': 2,
    'total_imp': 3,
    'rate_card': 4,
    'networks': 5,
    'invoice_num': 6,
    'start_point': 7,
    'bill_to': 8,
    'attention': 9,
    'address': 10,
    'state': 11,
    'contact': 12
}

turner_top = [
    f"truTV {start_date.strftime('%b %y')} Campaigns",
    f"Adult Swim {start_date.strftime('%b %y')} Campaigns",
    f"TBS {start_date.strftime('%b %y')} Campaigns",
    f"Boomerang {start_date.strftime('%b %y')} Campaigns",
    f"Cartoon Network {start_date.strftime('%b %y')} Campaigns",
    f"Cartoon Network ESP {start_date.strftime('%b %y')} Campaigns",
    f"CNN {start_date.strftime('%b %y')} Campaigns",
    f"HLN {start_date.strftime('%b %y')} Campaigns",
    f"TNT {start_date.strftime('%b %y')} Campaigns",
    f"March Madness {start_date.strftime('%b %y')} Campaigns"
]

# Total Impressions
query = f'''
SELECT os.PROGRAMMER, sum(os.IMPRESSIONS) FROM OPERATIONS.OPS_STAT_ALL os
WHERE os.EVENT_DATE >= '01-JAN-{today.strftime('%y')}'
AND os.EVENT_DATE <= '{end_date.strftime('%d-%b-%y').upper()}'
GROUP BY os.PROGRAMMER ORDER BY 1
'''
cur = cursor.execute(query)
impressions = cur.fetchall()

rate_card = {
    199999999: {
        'value_1': 0.95,
        'value_2': 1.05,
        'P4_1': 1.28,
        'P4_2':1.42
    },
    399999999: {
        'value_1': 0.84,
        'value_2': 1.00,
        'P4_1': 1.13,
        'P4_2': 1.35
    },
    599999999: {
        'value_1': 0.74,
        'value_2': 0.95,
        'P4_1': 0.99,
        'P4_2': 1.28
    },
    799999999: {
        'value_1': 0.63,
        'value_2': 0.89,
        'P4_1': 0.85,
        'P4_2': 1.21
    }, 
    1999999999: {
        'value_1': 0.53,
        'value_2': 0.84,
        'P4_1': 0.71,
        'P4_2': 1.13
    },
    2999999999: {
        'value_1': 0.45,
        'value_2': 0.79,
        'P4_1': 0.61,
        'P4_2': 1.06
    },
    3999999999: {
        'value_1': 0.43,
        'value_2': 0.76,
        'P4_1': 0.58,
        'P4_2': 1.03
    },
    4999999999: {
        'value_1': 0.41,
        'value_2': 0.73,
        'P4_1': 0.55,
        'P4_2': 0.99
    },
    99999999999: {
        'value_1': 0.41,
        'value_2': 0.73,
        'P4_1': 0.50,
        'P4_2': 0.94
    }
}
