import time
timer_s = time.time()

import openpyxl as xl
from config import *
from colorama import Fore


df_path = '../Canoe/INVOICES_2019/04/DATA/MRM_INVOICE_APR.xlsx'
df_wb = xl.load_workbook(df_path, data_only=True)

today_invoice = today.strftime('%m/%d/%Y')
today_filename = today.isoformat().replace('-', '')

def get_max_imp(imp):
    for rate in rate_card:
        if imp <= rate:
            return rate

def get_next_cpm(cpm):
    for r in range(17, 26):
        if invoice.cell(row=r, column=9).value == cpm:
            return invoice.cell(row=r+1, column=9).value

invoices = 0
for programmer in programmers:
    invoice_timer_s = time.time()
    try:
        # Setting Variables
        programmer_df = df_wb.get_sheet_by_name(programmers[programmer]['title'])

        programmer_wb = xl.load_workbook(programmers[programmer]['invoice_path'])
        invoice = programmer_wb.get_sheet_by_name('Invoice')

        df_length = programmer_df.max_row - 4
        start_point = programmers[programmer]['start_point']
        
        # Updatting Header
        invoice[invoice_map['date']] = today_invoice
        invoice[invoice_map['invoice_num']] = programmers[programmer]['invoice_num']
        invoice[invoice_map['period_start']] = start_date
        invoice[invoice_map['period_end']] = end_date
        invoice[invoice_map['previous_YTD_imp']] = programmers[programmer]['total_imp']

        # Setting Rate Card
        max_imp = get_max_imp(programmers[programmer]['total_imp'])
        next_cpm = get_next_cpm(programmers[programmer]['rate_card']['value'])

        # Pasting Data
        i=1
        imp_counter = programmers[programmer]['total_imp']
        r = start_point
        last_r = start_point + df_length
        while r < last_r:
            invoice.insert_rows(r)
            # Invoice Line
            invoice[f"B{r}"] = i
            # Data Frame
            for key in df_map:
                invoice[f"{invoice_map[key]}{r}"] = programmer_df[f"{df_map[key]}{i+3}"].value
            # Rate Card & Total
            imp_counter += invoice[f"{invoice_map['month_imp']}{r}"].value
            if imp_counter >= max_imp:
                split_imp = imp_counter - max_imp
                invoice[f"{invoice_map['month_imp']}{r}"] = programmer_df[f"{df_map['month_imp']}{i+3}"].value - split_imp
                invoice[f"{invoice_map['cpm']}{r}"] = programmers[programmer]['rate_card']['value']
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"

                r+=1
                invoice.insert_rows(r)
                programmers[programmer]['rate_card']['value'] = next_cpm
                invoice[f"{invoice_map['network']}{r}"] = invoice[f"{invoice_map['network']}{r-1}"].value
                invoice[f"{invoice_map['month_imp']}{r}"] = split_imp
                invoice[f"{invoice_map['cpm']}{r}"] = programmers[programmer]['rate_card']['value']
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"

                max_imp = get_max_imp(imp_counter)
                next_cpm = get_next_cpm(next_cpm)
                last_r+=1
            else:
                invoice[f"{invoice_map['cpm']}{r}"] = programmers[programmer]['rate_card']['value']
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"

            r+=1
            i+=1

        # Formatting Corrections
        for r in range(last_r + 2, last_r + 2 + len(programmers[programmer]['networks']) +25):
            # Update sub-totals
            if invoice.cell(row=r, column=8).value in programmers[programmer]['networks']:
                invoice.cell(row=r, column=9).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + df_length},H{r},{invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + df_length})"
                invoice.cell(row=r, column=11).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + df_length},H{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + df_length})"
            # Update total
            if invoice.cell(row=r, column=7).value == 'Total:':
                invoice.cell(row=r, column=9).value = f"=SUM({invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + df_length})"
                billed_imp = invoice.cell(row=r, column=9).value
                invoice.cell(row=r, column=11).value = f"=SUM({invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + df_length})"
                total_row = r
            # Update amount due
            if invoice.cell(row=r, column=10).value == 'Amount Due:':
                invoice.cell(row=r, column=11).value = f"=SUM({invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + df_length})"
                amount_due = invoice.cell(row=r, column=11).value
                # Update YTD impressions
                for i in range(17, 26):
                    if invoice.cell(row=i, column=9).value == programmers[programmer]['rate_card']['value']:
                        invoice.cell(row=i, column=10).value = f"=SUM({invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + df_length}) + {invoice_map['previous_YTD_imp']}"
        if programmer == 'FOX':
            invoice[f"{invoice_map['month_imp']}28"] = billed_imp
            invoice[f"{invoice_map['total']}28"] = amount_due

        elif programmer == 'TURNER':
            for r in range(28, 40):
                invoice.cell(row=r, column=9).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + df_length},E{r},{invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + df_length})"
                invoice.cell(row=r, column=11).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + df_length},E{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + df_length})"

        invoice_timer_e = time.time()
        programmer_wb.save(f'new_invoices/invoice-{programmer}-{today_filename}.xlsx')
        print(f"[{Fore.GREEN}SUCCESS{Fore.WHITE}] {programmers[programmer]['title']} invoice successfully generated in {round(invoice_timer_e - invoice_timer_s,3)}s")
        invoices += 1

    except:
        print(f"[{Fore.RED}ERROR{Fore.WHITE}] Fail to generate {programmers[programmer]['title']} invoice ...")

timer_e = time.time()
print(f'{invoices} invoices generated in {round(timer_e - timer_s, 3)}s')