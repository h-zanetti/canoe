import openpyxl as xl
import datetime as dt

today = dt.date.today()

time_period_start = input('Time period start (mm/dd/yyyy): ')
time_period_end = input('Time period end (mm/dd/yyyy): ')
df = input('Dataframe path: ')
revenue = input('Revenue path: ')

wb_df = xl.load_workbook(f'{df}.xlsx', data_only=True)
wb_revenue = xl.load_workbook(f'{revenue}.xlsx', data_only=True)

ws_revenue = wb_revenue.get_sheet_by_name('Revenue Calc 2019')
ws_impressions = wb_revenue.get_sheet_by_name('2019 Impressions')
ws_invoice_num = wb_revenue.get_sheet_by_name('Invoice Numbers')

invoice_numbers = {}
for n in range(3, 25):
    invoice_numbers[f"{ws_invoice_num[f'B{n}'].value}"] = ws_invoice_num[f'F{n}'].value

impressions = {}
for n in range(3, 23):
    impressions[f"{ws_impressions[f'B{n}'].value}"] = ws_impressions[f'F{n}'].value

revenue = {}
for n in range(5, 25):
    revenue[f"{ws_revenue[f'D{n}'].value}"] = round(ws_revenue[f'H{n}'].value)

# programmers = []
# for programmer in impressions:
#     programmers.append(programmer)

programmers = ['CW', 'Epix', 'Genius Brands', 'Kabillion' , 'Reelz' , 'Starz Entertainment', 'TV One']

# loop over programmers and create an invoice for each of them
# for programmer in programmers:
#     ws_invoice.title = programmer

# Map invoice locations and columns
invoice_map = {
# Header
    'date': 'L1', 'invoice_num': 'L2', 'period_start': 'D18', 'period_end': 'D19', 'previous_YTD_impressions': 'D22', 'YTD_impressions': 'K17',

# Body
    # All invoices from CANOE_INVOICE_APR_2019 starts at the column 27, for some weird reason CW starts at 28
    'data_starts': 'C28', 'data_ends': 'J28',
    # Data columns
    'campaign_id': 'C', 'campaign_name': 'D', 'network': 'E', 'start_date': 'F', 'end_date': 'G', 'campaign_goal': 'H', 'total_impressions': 'I',
    'month_impressions': 'J', 'cpm':  'K', 'total': 'L',
}

# Dataframe Column Map
df_map = {
    'campaign_id': 'A', 'campaign_name': 'B', 'network': 'C',
    'start_date': 'D', 'end_date': 'E',
    'campaign_goal': 'F', 'total_impressions': 'G', 'month_impressions': 'H'
}

cpm = [1.28, 1.13, 0.99, 0.85, 0.71, 0.61, 0.58, 0.55, 0.5]


file_name_date = today.isoformat().replace('-', '')
file_name = f'invoice-{programmers[0]}-{file_name_date}_P4.xlsx'
print(f'Creating {file_name}')

# Load CW dataframe
ws_CW_df = wb_df.get_sheet_by_name('CW')

# Load last CW invoice
wb_old_CW = xl.load_workbook('../INVOICE_MAR_2019/invoice-CW-20190401_P4.xlsx')
ws_old_CW = wb_old_CW.get_sheet_by_name('Invoice')

wb_old_CW_data = xl.load_workbook('../INVOICE_MAR_2019/invoice-CW-20190401_P4.xlsx', data_only=True)
ws_old_CW_data = wb_old_CW_data.get_sheet_by_name('Invoice')

# Change header
today.strftime('%m/%d/%Y')
ws_old_CW[invoice_map['date']] = today

ws_old_CW[invoice_map['invoice_num']] = invoice_numbers['CW']

ws_old_CW[invoice_map['previous_YTD_impressions']] = ws_old_CW_data[invoice_map['YTD_impressions']].value # incomplete

ws_old_CW[invoice_map['period_start']] = time_period_start
ws_old_CW[invoice_map['period_end']] = time_period_end

# Past data
i = 1
for n in range(28, 87+29):
    if i == ws_old_CW_data[f'B{n}'].value:
        for key in df_map:
                ws_old_CW[f"{invoice_map[key]}{n}"] = ws_CW_df[f"{df_map[key]}{n}"].value
    else:
        # ws_old_CW.insert_rows(1)
        ws_old_CW[f'B{n}'] = f'=B{n - 1}+1'
        for key in df_map:
                ws_old_CW[f"{invoice_map[key]}{n}"] = ws_CW_df[f"{df_map[key]}{n}"].value
    i += 1