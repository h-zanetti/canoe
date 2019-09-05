# incomlete
import openpyxl as xl

new_wb = xl.Workbook()

num_sheets = int(input("Number of sheets: "))
for n in range(0, num_sheets):
    directory = input("File directory: ")
    file_name = input("File name: ")
    wb = xl.load_workbook(f"{directory}/{file_name}")

    sheet = input("Sheet title: ")
    ws = wb.get_sheet_by_name(sheet)
    new_ws = new_wb.create_sheet(title)

file_path = input("Save file as: ")
wb.save(file_path)