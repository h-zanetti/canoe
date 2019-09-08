import time
timer_s = time.time()

import openpyxl as xl
from colorama import Fore
from config import *

# Dataframe
df_path = f"../data/{start_date.strftime('%b').upper()}_INVOICE.xlsx"
df_wb = xl.load_workbook(df_path, data_only=True)

# Date Formatting
today_invoice = today.strftime('%m/%d/%Y')
filename_date = start_date.strftime('%b_%Y').upper()

invoices = 0
programmer_wb = xl.load_workbook('../plain_invoices/PLAIN_INVOICES.xlsx')
for programmer in programmers:
    invoice_timer_s = time.time()
    
    # try:
    # Setting Variables
    programmer_df = df_wb.get_sheet_by_name(programmer[db_map['title']])

    df_length = programmer_df.max_row - 4
    start_point = programmer[db_map['start_point']]

    networks_str = programmer[db_map['networks']]
    networks_list = networks_str.split(', ')

    invoice = programmer_wb.get_sheet_by_name(programmer[db_map['abbreviation']])

    # Updatting Header
    invoice[invoice_map['date']] = today_invoice
    basic_font(invoice[invoice_map['date']])
    alignment(invoice[invoice_map['date']], 'right')

    query = "SELECT MAX(invoice_num) FROM invoice_generator_info"
    cursor.execute(query)
    invoice_num = cursor.fetchone()
    if invoices == 0:
        invoice_num = invoice_num[0] + 4
    else:
        invoice_num = invoice_num[0] + 1
    # invoice_num = invoice_num[0] + 1
    invoice[invoice_map['invoice_num']] = invoice_num
    basic_font(invoice[invoice_map['invoice_num']])
    alignment(invoice[invoice_map['invoice_num']], 'right')

    invoice[invoice_map['period_start']] = start_date
    invoice[invoice_map['period_end']] = end_date
    if programmer[db_map['abbreviation']] == 'DISCOVERY':
        invoice['D24'] = programmer[db_map['total_imp']]
    else:
        invoice[invoice_map['previous_YTD_imp']] = programmer[db_map['total_imp']]

    # Setting Rate Card Variables
    current_cpm = programmer[db_map['rate_card']]
    max_imp = get_max_imp(programmer[db_map['total_imp']])

    # Pasting Data
    i=1
    imp_counter = programmer[db_map['total_imp']]
    backfill_impressions = 0
    r = start_point
    last_r = start_point + df_length
    while r <= last_r:
        invoice.insert_rows(r)
        # Line Number
        cell = invoice[f"B{r}"]
        cell.value = i
        cell.number_format = '000'
        
        # Data Frame
        for key in df_map:
            # Date Formatting
            if key == 'start_date' or key == 'end_date':
                cell = invoice[f"{invoice_map[key]}{r}"]
                cell.value = programmer_df[f"{df_map[key]}{i+3}"].value
                cell.number_format = 'MM/DD/YYYY'
            else:
                invoice[f"{invoice_map[key]}{r}"] = programmer_df[f"{df_map[key]}{i+3}"].value
        
        # Rate Card & Total
        if invoice[f"{invoice_map['network']}{r}"].value in backfill['keys'] or invoice[f"{invoice_map['campaign_name']}{r}"].value in backfill['keys']:
            invoice[f"{invoice_map['cpm']}{r}"] = 0
            invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"
            currency_format(invoice[f"{invoice_map['cpm']}{r}"])
            currency_format(invoice[f"{invoice_map['total']}{r}"])
        else:
            imp_counter += invoice[f"{invoice_map['month_imp']}{r}"].value

            if imp_counter >= max_imp:
                split_imp = imp_counter - max_imp
                invoice[f"{invoice_map['month_imp']}{r}"] = programmer_df[f"{df_map['month_imp']}{i+3}"].value - split_imp
                invoice[f"{invoice_map['cpm']}{r}"] = current_cpm
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"
                invoice_font(r, invoice)
                currency_format(invoice[f"{invoice_map['cpm']}{r}"])
                currency_format(invoice[f"{invoice_map['total']}{r}"])

                r+=1
                invoice.insert_rows(r)
                current_cpm = get_next_cpm(current_cpm, invoice)
                invoice[f"{invoice_map['network']}{r}"] = invoice[f"{invoice_map['network']}{r-1}"].value
                invoice[f"{invoice_map['month_imp']}{r}"] = split_imp
                invoice[f"{invoice_map['cpm']}{r}"] = current_cpm
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"
                currency_format(invoice[f"{invoice_map['cpm']}{r}"])
                currency_format(invoice[f"{invoice_map['total']}{r}"])

                max_imp = get_max_imp(imp_counter)
                last_r+=1
            else:
                invoice[f"{invoice_map['cpm']}{r}"] = current_cpm
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"
                currency_format(invoice[f"{invoice_map['cpm']}{r}"])
                currency_format(invoice[f"{invoice_map['total']}{r}"])

        invoice_font(r, invoice)
        r+=1
        i+=1

    # Formatting Corrections
    invoice_max = invoice.max_row + 1
    for r in range(last_r + 2, invoice_max):
        # Update sub-totals by network
        if invoice[f"{invoice_map['sub_networks']}{r}"].value in networks_list or invoice[f"{invoice_map['sub_networks']}{r}"].value == "Backfill Networks":
            invoice[f"{invoice_map['month_imp']}{r}"] = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},{invoice_map['sub_networks']}{r},{invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + i})"
            invoice[f"{invoice_map['total']}{r}"] = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},{invoice_map['sub_networks']}{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            currency_format(invoice[f"{invoice_map['total']}{r}"])

        if invoice[f"{invoice_map['foot_networks']}{r}"].value in networks_list:
            invoice[f"{invoice_map['total']}{r}"] = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},{invoice_map['foot_networks']}{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            currency_format(invoice[f"{invoice_map['total']}{r}"])

        # Update total
        if invoice[f"{invoice_map['foot_total']}{r}"].value == 'Total:':
            invoice[f"{invoice_map['month_imp']}{r}"] = imp_counter - programmer[db_map['total_imp']]
            local_imp = invoice[f"{invoice_map['month_imp']}{r}"].value

            invoice[f"{invoice_map['total']}{r}"] = f"=SUM({invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            currency_format(invoice[f"{invoice_map['total']}{r}"])
            
        # Update amount due
        if invoice[f"{invoice_map['amount_due']}{r}"].value == 'Amount Due:':
            invoice[f"{invoice_map['total']}{r}"] = f"=SUM({invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            amount_due = invoice[f"{invoice_map['total']}{r}"].value
            # Update YTD impressions
            for row in range(17, 26):
                if round(invoice[f"{invoice_map['rate_cards']}{row}"].value, 2) == current_cpm:
                    invoice[f"{invoice_map['YTD_imp']}{row}"] = imp_counter
                    alignment(invoice[f"{invoice_map['YTD_imp']}{row}"], 'center')
                    if imp_counter >= 1_000_000_000:
                        invoice[f"{invoice_map['YTD_imp']}{row}"].number_format = '#0.00,,, "B"'
                    else:
                        invoice[f"{invoice_map['YTD_imp']}{row}"].number_format = '#0.00,, "M"'
                    for col in range(8, 11):
                        cell = invoice.cell(row=row, column=col)
                        cell.font = Font(name='Calibri', size=12, bold=True)
                        cell.fill = PatternFill('solid', fgColor='FFFF99')
                    break
                        
    # Handling Exceptions
    if programmer[db_map['abbreviation']] == 'FOX':
        invoice[f"{invoice_map['month_imp']}28"] = local_imp
        invoice[f"{invoice_map['total']}28"] = amount_due
        invoice[f"{invoice_map['campaign_name']}28"] = f"{start_date.strftime('%b %Y').upper()} Campaigns"

    elif programmer[db_map['abbreviation']] == 'TURNER':
        k=0
        for r in range(28, 38):
            invoice[f"{invoice_map['campaign_name']}{r}"] = turner_top[k]
            invoice[f"I{r}"] = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},{invoice_map['network']}{r},{invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + i})"
            invoice[f"{invoice_map['total']}{r}"] = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},{invoice_map['network']}{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            k+=1
        invoice[f"{invoice_map['total']}39"] = amount_due

    # Checking and Updating DB
    if imp_counter == impressions[programmer[db_map['id']]-1][1]:
        print(f"[{Fore.GREEN}OK{Fore.WHITE}] Total impressions for {programmer[db_map['abbreviation']]} match")
    else:
        print(f"[{Fore.YELLOW}ALERT{Fore.WHITE}] Total impressions for {programmer[db_map['abbreviation']]} doesn't match")
    
    # Print Area
    invoice.print_area = f'A1:K{invoice.max_row}'

    programmer_wb.save(f"../new_invoices/INVOICES_{filename_date}.xlsx")

    query = update(programmer[db_map['id']], imp_counter, current_cpm, invoice_num)
    cursor.execute(query)
    db.commit()

    invoice_timer_e = time.time()

    print(f"[{Fore.GREEN}SUCCESS{Fore.WHITE}] {programmer[db_map['abbreviation']]} invoice successfully generated in {round(invoice_timer_e - invoice_timer_s,3)}s")
    invoices += 1

    # except Exception as e:
    #     print(f"[{Fore.RED}ERROR{Fore.WHITE}] Fail to generate {programmer[db_map['title']]} invoice or update database...")
    #     print(e)

timer_e = time.time()
print(f'{invoices} invoices generated in {round(timer_e - timer_s, 3)}s')