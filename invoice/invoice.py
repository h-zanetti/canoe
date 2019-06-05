import time
timer_s = time.time()

import openpyxl as xl
from openpyxl.styles import Font, PatternFill
from colorama import Fore
from config import *

# Styling Classes
font = Font(name='Calibri', size=12)
fill = PatternFill()
def invoice_font(row):
    for col in range(2, 12):
        cell = invoice.cell(row=row, column=col)
        cell.font = Font(name='Calibri', size=12)

# Dataframe
df_path = f"data/INVOICE_{start_date.strftime('%b').upper()}.xlsx"
df_wb = xl.load_workbook(df_path, data_only=True)

# Date Formatting
today_invoice = today.strftime('%m/%d/%Y')
filename_date = start_date.strftime('%b_%Y').upper()

def get_max_imp(imp):
    for rate in rate_card:
        if imp <= rate:
            return rate

def get_next_cpm(cpm, invoice):
    for r in range(17, 26):
        if round(invoice[f"{invoice_map['rate_cards']}{r}"].value, 2) == cpm:
            return round(invoice[f"{invoice_map['rate_cards']}{r+1}"].value, 2)

def update(abr, total_imp, rate_card, invoice_num):
        dic = {
        'total_imp': total_imp,
        'rate_card': rate_card,
        'invoice_num': invoice_num,
        }

        update = f'UPDATE invoice_generator_info SET '
        for key in dic:
            update += f'{key} = {dic[key]}, '
        update = update[:-2]
        update += f" WHERE abbreviation = '{abr}'"
       
        return update

invoices = 0
db_map = {
    'id': 0,
    'abbreviation': 1,
    'title': 2,
    'total_imp': 3,
    'rate_card': 4,
    'networks': 5,
    'invoice_num': 6,
    'start_point': 7,
    'bill_to': 8,
    'attention': 9,
    'address': 10,
    'state': 11,
    'contact': 12
}
programmer_wb = xl.load_workbook('plain_invoices/PLAIN_INVOICES.xlsx')
for programmer in programmers:
    invoice_timer_s = time.time()
    
    try:
        # Setting Variables
        programmer_df = df_wb.get_sheet_by_name(programmer[db_map['title']])

        df_length = programmer_df.max_row - 4
        start_point = programmer[db_map['start_point']]

        networks_str = programmer[db_map['networks']]
        networks_list = networks_str.split(', ')

        invoice = programmer_wb.get_sheet_by_name(programmer[db_map['abbreviation']])
        
        # Updatting Header
        invoice[invoice_map['date']] = today_invoice

        query = "SELECT MAX(invoice_num) FROM invoice_generator_info"
        cursor.execute(query)
        invoice_num = cursor.fetchone()
        invoice_num = invoice_num[0] + 1
        invoice[invoice_map['invoice_num']] = invoice_num

        invoice[invoice_map['period_start']] = start_date
        invoice[invoice_map['period_end']] = end_date
        if programmer[db_map['abbreviation']] == 'DISCOVERY':
            invoice['D24'] = programmer[db_map['total_imp']]
        else:
            invoice[invoice_map['previous_YTD_imp']] = programmer[db_map['total_imp']]

        # Setting Rate Card Variables
        current_cpm = programmer[db_map['rate_card']]
        max_imp = get_max_imp(programmer[db_map['total_imp']])

        # Pasting Data
        i=1
        imp_counter = programmer[db_map['total_imp']]
        r = start_point
        last_r = start_point + df_length
        while r <= last_r:
            invoice.insert_rows(r)
            # Line Number
            cell = invoice[f"B{r}"]
            cell.value = i
            cell.number_format = '000'
            # Data Frame
            for key in df_map:
                if key == 'start_date' or key == 'end_date':
                    cell = invoice[f"{invoice_map[key]}{r}"]
                    cell.value = programmer_df[f"{df_map[key]}{i+3}"].value
                    cell.number_format = 'MM/DD/YYYY'
                else:
                    invoice[f"{invoice_map[key]}{r}"] = programmer_df[f"{df_map[key]}{i+3}"].value
            # Rate Card & Total
            imp_counter += invoice[f"{invoice_map['month_imp']}{r}"].value
            if imp_counter >= max_imp:
                split_imp = imp_counter - max_imp
                invoice[f"{invoice_map['month_imp']}{r}"] = programmer_df[f"{df_map['month_imp']}{i+3}"].value - split_imp
                invoice[f"{invoice_map['cpm']}{r}"] = current_cpm
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"
                invoice_font(r)

                r+=1
                invoice.insert_rows(r)
                current_cpm = get_next_cpm(current_cpm, invoice)
                invoice[f"{invoice_map['network']}{r}"] = invoice[f"{invoice_map['network']}{r-1}"].value
                invoice[f"{invoice_map['month_imp']}{r}"] = split_imp
                invoice[f"{invoice_map['cpm']}{r}"] = current_cpm
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"

                max_imp = get_max_imp(imp_counter)
                last_r+=1
            else:
                invoice[f"{invoice_map['cpm']}{r}"] = current_cpm
                invoice[f"{invoice_map['total']}{r}"] = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"

            invoice_font(r)
            r+=1
            i+=1

        # Formatting Corrections
        invoice_max = invoice.max_row + 1
        for r in range(last_r + 2, invoice_max):
            # Update sub-totals
            if invoice.cell(row=r, column=8).value in networks_list:
                invoice.cell(row=r, column=9).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},H{r},{invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + i})"
                invoice.cell(row=r, column=11).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},H{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            if invoice.cell(row=r, column=10).value in networks_list:
                invoice.cell(row=r, column=11).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},J{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            # Update total
            if invoice.cell(row=r, column=7).value == 'Total:' or invoice.cell(row=r, column=7).value == "TOTAL:":
                invoice.cell(row=r, column=9).value = f"=SUM({invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + i})"
                local_imp = invoice.cell(row=r, column=9).value

                invoice.cell(row=r, column=11).value = f"=SUM({invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
            # Update amount due
            if invoice.cell(row=r, column=10).value == 'Amount Due:':
                invoice.cell(row=r, column=11).value = f"=SUM({invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"
                amount_due = invoice.cell(row=r, column=11).value
                # Update YTD impressions
                for row in range(17, 26):
                    if round(invoice.cell(row=row, column=9).value, 2) == current_cpm:
                        invoice.cell(row=row, column=10).value = f"=SUM({invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + i}) + {invoice_map['previous_YTD_imp']}"
                        for col in range(7, 12):
                            cell = invoice.cell(row=row, column=col)
                            cell.font = Font(name='Calibri', size=12, bold=True)
                            cell.fill = PatternFill('solid', fgColor='FFFF99')
                        break
                            
        # Handling Exceptions
        if programmer[db_map['abbreviation']] == 'FOX':
            invoice[f"{invoice_map['month_imp']}28"] = local_imp
            invoice[f"{invoice_map['total']}28"] = amount_due

        elif programmer[db_map['abbreviation']] == 'TURNER':
            for r in range(28, 40):
                invoice.cell(row=r, column=9).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},E{r},{invoice_map['month_imp']}{start_point}:{invoice_map['month_imp']}{start_point + i})"
                invoice.cell(row=r, column=11).value = f"=SUMIF({invoice_map['network']}{start_point}:{invoice_map['network']}{start_point + i},E{r},{invoice_map['total']}{start_point}:{invoice_map['total']}{start_point + i})"

        # Checking and Updating DB
        if imp_counter == impressions[programmer[db_map['id']]-1][1]:
            print(f"[{Fore.GREEN}OK{Fore.WHITE}] Total impressions for {programmer[db_map['abbreviation']]} match")
        else:
            print(f"[{Fore.YELLOW}ALERT{Fore.WHITE}] Total impressions for {programmer[db_map['abbreviation']]} doesn't match")
        
        programmer_wb.save(f"new_invoices/INVOICES_{filename_date}.xlsx")

        query = update(programmer[db_map['abbreviation']], imp_counter, current_cpm, invoice_num)
        cursor.execute(query)
        db.commit()

        invoice_timer_e = time.time()

        print(f"[{Fore.GREEN}SUCCESS{Fore.WHITE}] {programmer[db_map['abbreviation']]} invoice successfully generated in {round(invoice_timer_e - invoice_timer_s,3)}s")
        invoices += 1

    except Exception as e:
        print(f"[{Fore.RED}ERROR{Fore.WHITE}] Fail to generate {programmer[db_map['title']]} invoice or update database...")
        print(e)

timer_e = time.time()
print(f'{invoices} invoices generated in {round(timer_e - timer_s, 3)}s')