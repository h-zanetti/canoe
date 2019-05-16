import openpyxl as xl
import datetime as dt
from colorama import Fore

today = dt.date.today()

# Immutable Data
invoice_map = {
    # Header
    'date': 'K1', 'invoice_num': 'K2', 'period_start': 'D18', 'period_end': 'D19', 'previous_YTD_imp': 'D22',
    # Body
    'YTD_imp': 'J', 'campaign_id': 'C', 'campaign_name': 'D', 'network': 'E', 'start_date': 'F', 'end_date': 'G', 'total_imp': 'H',
    'month_imp': 'I', 'cpm':  'J', 'total': 'K',
}
df_map = {
    'campaign_id': 'A', 'campaign_name': 'B', 'network': 'C',
    'start_date': 'D', 'end_date': 'E',
    'total_imp': 'F', 'month_imp': 'G'
}
programmers = {
    'Viacom': 'P4',
    # 'NBC Universal': None,
    # 'ABC Disnay': 'P4',
    # 'AMC Networks': 'P4',
    # 'CBS Corporation': 'P4',
    # 'CW': 'P4',
    # 'Epix': None,
    # 'Genius Brands': None,
    # 'Kabillion': None,
    # 'Reelz': None,
    # 'Starz Entertainment': None,
    # 'TV One': None
}
cpms = {
    range(0, 2 * 10**8): 1.28,
    range(2 * 10**8, 4 * 10**8): 1.13,
    range(4 * 10**8, 6 * 10**8): 0.99,
    range(6 * 10**8, 8 * 10**8): 0.85,
    range(8 * 10**8, 2 * 10**9): 0.71,
    range(2 * 10**9, 3 * 10**9): 0.61,
    range(3 * 10**9, 4 * 10**9): 0.58,
    range(4 * 10**9, 5 * 10**9): 0.55,
}


# Inputs
time_period_start = '04/01/2019' #input('Time period start (mm/dd/yyyy): ')
time_period_end = '04/30/2019' #input('Time period end (mm/dd/yyyy): ')

df_path = '../Canoe/INVOICES_2019/04/DATA/MRM_INVOICE_APR.xlsx' #input('Dataframe path: ')
revenue_path = '../Canoe/INVOICES_2019/Revenue(2019).xlsx' #input('Revenue path: ')

print("Collecting data...")
# Revenue Workbook
wb_revenue = xl.load_workbook(f'{revenue_path}', data_only=True)

# Revenue Sheets
ws_revenue = wb_revenue.get_sheet_by_name('Revenue Calc 2019')
ws_impressions = wb_revenue.get_sheet_by_name('2019 Impressions')
ws_invoice_num = wb_revenue.get_sheet_by_name('Invoice Numbers')

# Creating Dictionaries
invoice_numbers = {}
for n in range(3, 25):
    invoice_numbers[f"{ws_invoice_num[f'B{n}'].value}"] = ws_invoice_num[f'F{n}'].value

month_impressions = {}
for n in range(3, 23):
    month_impressions[f"{ws_impressions[f'B{n}'].value}"] = ws_impressions[f'F{n}'].value

impressions = {}
for r in range(3, 23):
    if ws_impressions.cell(row=r, column=2).value not in impressions:
        impressions[ws_impressions.cell(row=r, column=2).value] = 0
        for c in range(3,7):
            impressions[ws_impressions.cell(row=r, column=2).value] += ws_impressions.cell(row=r, column=c).value

revenue = {}
for n in range(5, 25):
    revenue[f"{ws_revenue[f'D{n}'].value}"] = round(ws_revenue[f'H{n}'].value)

programmer_cpm = {}
for programmer in impressions:
    for cpm in cpms:
        if impressions[programmer] in cpm:
            programmer_cpm[programmer] = cpms[cpm]

wb_df = xl.load_workbook(f'{df_path}', data_only=True)

print(Fore.WHITE + '[' + Fore.GREEN + "OK" + Fore.WHITE + '] Data successfully collected.')

for programmer in programmers:
    print(f"Creating invoice for {programmer}...")
    invoice_start = int(input("Column number where the data will start to be placed: "))
    df_length = int(input('Dataframe length: '))

    invoice_map['data_starts'] = f"C{invoice_start}"
    invoice_map['data_ends'] = f"J{invoice_start}"

    # try:
    ws_df = wb_df.get_sheet_by_name(programmer)

    old_invoice_path = input(f"Path of an old invoice for {programmer}: ")
    marketplace = int(input('How many marketplace fields are in the invoice? '))

    file_name_date = today.isoformat().replace('-', '')
    file_name_programmer = programmer.replace(' ', '_')
    if programmers[programmer] == 'P4':
        file_name = f'invoice-{file_name_programmer}-{file_name_date}_P4.xlsx'
    else:
        file_name = f'invoice-{file_name_programmer}-{file_name_date}.xlsx'

    print(f'Creating file {file_name}')

    wb_invoice = xl.load_workbook(old_invoice_path)
    ws_invoice = wb_invoice.get_sheet_by_name('Invoice')

    wb_invoice_data = xl.load_workbook(old_invoice_path, data_only=True)
    ws_invoice_data = wb_invoice_data.get_sheet_by_name('Invoice')

    # Update header
    ws_invoice[invoice_map['date']] = today.strftime('%m/%d/%Y')

    ws_invoice[invoice_map['invoice_num']] = invoice_numbers[programmer]

    for r in range(17, 26):
        if ws_invoice_data.cell(row=r, column=10).value != None:
            current_cpm = round(ws_invoice_data.cell(row=r, column=9).value, 2)
            ws_invoice[invoice_map['previous_YTD_imp']] = ws_invoice_data.cell(row=r, column=10).value
    
    # ALERT: CHECK RATE CARD
    cpm_change = False
    if current_cpm != programmer_cpm[programmer]:
        cpm_change = True
        print(f"{Fore.WHITE}[{Fore.YELLOW}ALERT{Fore.WHITE}] CHECK RATE CARD")

    ws_invoice[invoice_map['period_start']] = time_period_start
    ws_invoice[invoice_map['period_end']] = time_period_end


    # Update data
    i = 1
    networks = {}
    imp_counter = ws_invoice[f"{invoice_map['previous_YTD_imp']}"].value
    for r in range(invoice_start, invoice_start + df_length):
        # Line number
        if i > 1:
            ws_invoice[f'B{r}'] = f'=B{r - 1}+1'
        else:
            ws_invoice[f'B{r}'] = 1
        # Paste data
        if ws_invoice_data[f"{invoice_map['campaign_id']}{r}"].value != 'NA' and \
           ws_invoice_data[f"{invoice_map['network']}{r}"].value != None:
            for key in df_map:
                ws_invoice[f"{invoice_map[key]}{r}"] = ws_df[f"{df_map[key]}{i+3}"].value
        else:
            ws_invoice.insert_rows(r)
            for key in df_map:
                ws_invoice[f"{invoice_map[key]}{r}"] = ws_df[f"{df_map[key]}{i+3}"].value
        
        # Rate card
        imp_counter += ws_invoice[f"{invoice_map['month_imp']}{r}"].value
        if cpm_change == True:
            for cpm in cpms:
                if imp_counter in cpm and cpms[cpm] != current_cpm:
                    current_cpm = cpms[cpm]
                    ws_invoice[f"{invoice_map['cpm']}{r}"].value = current_cpm
                else:
                    ws_invoice[f"{invoice_map['cpm']}{r}"].value = current_cpm
        else:
            ws_invoice[f"{invoice_map['cpm']}{r}"].value = current_cpm
            
        # Total
        ws_invoice.cell(row=r, column=11).value = f"=ROUND({invoice_map['month_imp']}{r}*({invoice_map['cpm']}{r}/1000),2)"
        
        # Impressions per networks
        if ws_df[f"{df_map['network']}{i+3}"].value in networks:
            networks[ws_df[f"{df_map['network']}{i+3}"].value] += ws_df[f"{df_map['month_imp']}{i+3}"].value
        else:
            networks[ws_df[f"{df_map['network']}{i+3}"].value] = ws_df[f"{df_map['month_imp']}{i+3}"].value
        
        i += 1

    # Formatting corrections
    max_row = ws_invoice.max_row
    for r in range(invoice_start + df_length, invoice_start + df_length + len(networks) + marketplace + 25):
        # Update sub-totals
        if ws_invoice.cell(row=r, column=8).value in networks:
            ws_invoice.cell(row=r, column=9).value = networks[ws_invoice.cell(row=r, column=8).value]
            ws_invoice.cell(row=r, column=11).value = networks[ws_invoice.cell(row=r, column=8).value] * (current_cpm / 1000)
        # Update total
        if ws_invoice.cell(row=r, column=7).value == 'Total:':
            ws_invoice.cell(row=r, column=9).value = f"=SUM({invoice_map['month_imp']}{invoice_start}:{invoice_map['month_imp']}{invoice_start + df_length + marketplace - 1})"
            ws_invoice.cell(row=r, column=11).value = f"=SUM({invoice_map['total']}{invoice_start}:{invoice_map['total']}{invoice_start + df_length + marketplace - 1})"
        # Update amount due
        if ws_invoice.cell(row=r, column=10).value == 'Amount Due:':
            ws_invoice.cell(row=r, column=11).value = f"=SUM({invoice_map['total']}{invoice_start}:{invoice_map['total']}{invoice_start + df_length + marketplace - 1})"    
            amount_due_row = r
            for i in range(17, 26):
                if ws_invoice.cell(row=i, column=10).value != None:
                    ws_invoice.cell(row=i, column=10).value = f"={invoice_map['previous_YTD_imp']}+{ws_invoice.cell(row=amount_due_row, column=11).value}"

    wb_invoice.save(f"new_invoices/{file_name}")
    print(Fore.WHITE + '[' + Fore.GREEN + "SUCCESS" + Fore.WHITE + f"] {programmer} invoice created")

    # except:
    #     print('[' + Fore.RED + "ERROR" + Fore.WHITE + f'] Fail to generate {programmer} invoice')
